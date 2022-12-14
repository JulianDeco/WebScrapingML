import os

#Necesario para el scrapping
from bs4 import BeautifulSoup
import requests
import time

#Necesario para ordenar el diccionario
import operator

#Necesario para manejar hoja de calculo
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.dimensions import ColumnDimension

#Importamos la función para crear logs
from crear_logs import crear_log

search = input('Introduzca lo que quiera buscar: ')
#URL donde solicitamos el html
url = requests.get(f'https://listado.mercadolibre.com.ar/{search}')
soup = BeautifulSoup(url.content, "html.parser")


# Variables donde almacenamos los datos que necesitamos
resultados_titulos = soup.findAll('h2', {'class':'ui-search-item__title'})
# resultados_precios = soup.findAll('span', {'class':'price-tag-text-sr-only'})
resultados_precios = soup.findAll('div', {'class':'ui-search-price__second-line'})
#Creacion diccionario
diccionario_datos = {}
#Bucle FOR para obtener resultados de a uno y trabajarlos individualmente
#tambien almacenamos los datos en un diccionario
for i in range(0,(len(resultados_precios))):
    
    res_precios_texto = resultados_precios[i].text
    res_precios_texto_sin_coma = res_precios_texto.split()
    
    res_titulos_texto = resultados_titulos[i]
    
    diccionario_datos[res_titulos_texto.text] = int(res_precios_texto_sin_coma[0])


#Ordenamos el diccionario de menor a mayor en relacion a la key
#Esto nos devuelve una lista
diccionario_datos_ordenados = sorted(diccionario_datos.items(), key=operator.itemgetter(1))

print(diccionario_datos_ordenados)

#Comprobamos si existe el archivo
if os.path.exists('precios_ml.xlsx'):
    print('Ya existe un archivo con este nombre')
    
else:

    #Creamos un libro
    wb = Workbook()
    
    #Creamos el archivo
    ws = wb.active    
    ws.title = 'ARTICULOS Y PRECIOS'
    wb.save('precios_ml.xlsx')
    print('Archivo creado')

#Nombre archivo
file = "./precios_ml.xlsx"

#Cargamos el archivo
wb = load_workbook(file)

#Usamos el archivo abierto
ws = wb.active    



# #Creamos la hoja, con el 0 indicamos que se cree primero


sheet = input("Nombre de la página a crear: ")
ws = wb.create_sheet(f'{sheet}',0)    


#Creamos las columnas
ws['A1'] = 'ARTICULO'
ws['B1'] = 'PRECIO'

cont = 2
#Bucle para escribir la hoja
for art_precios in diccionario_datos_ordenados:
    
    ws[f'A{cont}'] = art_precios[0]
    ws[f'B{cont}'] = art_precios[1]
    
    ws[f'B{cont}'].number_format = '_-[$$-es-AR] * #,##0.00_-;-[$$-es-AR] * #,##0.00_-;_-[$$-es-AR] * "-"??_-;_-@_-' 
    
    cont += 1
    
ws.column_dimensions['A'].width = 90
ws.column_dimensions['B'].width = 30    

    


#Guardamos el archivo
wb.save(file)

crear_log(len(diccionario_datos_ordenados), search)


if __name__ == '__main__':
    print('Busqueda finalizada y ejecutada con éxito')
    #Imprimimos el diccinario por prueba
    # for a in diccionario_datos_ordenados:
    #     print(a)


