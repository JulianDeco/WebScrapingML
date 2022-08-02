import os

#Necesario para el scrapping
from bs4 import BeautifulSoup
import requests
import time

#Necesario para ordenar el diccionario
import operator

#Necesario para manejar hoja de calculo
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.dimensions import ColumnDimension

from crear_logs import crear_log


print("1-Calzado\n2-Indumentaria\n13-Accesorios")
search = input('Elige una opción ')

marca_DS = input('Las opciones son "fluid", "sport78", "dionysos" ,"blast" ,"sale-fin-de-temporada", "tienda-futbol-78" o "dia-del-nino-digital-sport": ')

#URL donde solicitamos el html
url = requests.get(f'https://www.digitalsport.com.ar/{marca_DS}/prods/?category[1]={search}')
print(url)
soup = BeautifulSoup(url.content, "html.parser")
#Variables donde almacenamos los datos que necesitamos
resultados_titulos = soup.findAll('h3')
resultados_precios = soup.findAll('div', {'class':'precio'})
resultados_marca = soup.findAll('div', {'class':'brand'})

#Creacion diccionario
diccionario_datos = {}

for i in range(0,(len(resultados_titulos))):
    res_titulos_texto = resultados_titulos[i]
    res_precios_texto = resultados_precios[i]
    res_marca_texto = resultados_marca[i]
    
    res_precios_sin_coma = res_precios_texto.text
    res_precios_texto_sin_coma = int(res_precios_sin_coma.replace('$', '')) 
    diccionario_datos[f'{(res_marca_texto.text).upper()} - {res_titulos_texto.text}'] = res_precios_texto_sin_coma

print(diccionario_datos)

diccionario_datos_ordenados = sorted(diccionario_datos.items(), key=operator.itemgetter(1))

#Comprobamos si existe el archivo
if os.path.exists('precios_ml.xlsx'):
    print('Ya existe un archivo con este nombre')
    
else:

    #Creamos un libro
    wb = Workbook()
    
    #Creamos el archivo
    ws = wb.active    
    ws.title = 'ARTICULOS Y PRECIOS'
    wb.save('precios_ml.xlsx')
    print('Archivo creado')

#Nombre archivo
file = "./precios_ml.xlsx"

#Cargamos el archivo
wb = load_workbook(file)

#Usamos el archivo abierto
ws = wb.active    



# #Creamos la hoja, con el 0 indicamos que se cree primero


sheet = input("Nombre de la página a crear: ")
ws = wb.create_sheet(f'{sheet}',0)    


#Creamos las columnas
ws['A1'] = 'ARTICULO'
ws['B1'] = 'PRECIO'

cont = 2
#Bucle para escribir la hoja
for art_precios in diccionario_datos_ordenados:
    
    ws[f'A{cont}'] = art_precios[0]
    ws[f'B{cont}'] = art_precios[1]
    
    ws[f'B{cont}'].number_format = '_-[$$-es-AR] * #,##0.00_-;-[$$-es-AR] * #,##0.00_-;_-[$$-es-AR] * "-"??_-;_-@_-' 
    
    cont += 1
    
ws.column_dimensions['A'].width = 90
ws.column_dimensions['B'].width = 30    

    


#Guardamos el archivo
wb.save(file)

crear_log(len(diccionario_datos_ordenados), search)